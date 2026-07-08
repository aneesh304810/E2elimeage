"""
Legacy end-to-end lineage connector.

Reads two sheets from the legacy lineage workbook:
  * a lineage/schema sheet: one row per DWH target column, mapping it back
    through STG2 -> STG1 -> SRC with the transformation at each hop and a
    Lineage_Status.
  * a "proof" sheet: sample values per field per stage (DWH / STG2 / STG1),
    used to eyeball data variance across stages.

The USER_DEFINED_ATTRIBUTE_CLOB column holds a JSON blob like
{"UD_1":"...","UD_7":"..."}. Each UD_n is exploded into its own lineage +
proof rows (is_ud='Y', ud_key='UD_7'), while the original CLOB field is ALSO
kept as a normal field, so both the flat and exploded views are available.

Env:
  CP_LEGACY_LINEAGE_XLSX  path to the workbook
  (optional) CP_LEGACY_LINEAGE_SHEET / CP_LEGACY_PROOF_SHEET  sheet names
"""
from __future__ import annotations
import os, json, logging
from openpyxl import load_workbook

log = logging.getLogger("cp.legacy_lineage")

UD_CLOB_COL = "USER_DEFINED_ATTRIBUTE_CLOB"

# lineage sheet header -> table column
LINEAGE_MAP = {
    "DWH_Target_Table": "dwh_target_table",
    "DWH_Target_Column": "dwh_target_column",
    "DWH_Type": "dwh_type", "DWH_Length": "dwh_length", "DWH_Precision": "dwh_precision",
    "STG2_Source_Table": "stg2_source_table", "STG2_Source_Column": "stg2_source_column",
    "STG2_to_DWH_Transformation": "stg2_to_dwh_transform",
    "STG2_Type": "stg2_type", "STG2_Length": "stg2_length", "STG2_Precision": "stg2_precision",
    "STG1_Source_Table": "stg1_source_table", "STG1_Source_Column": "stg1_source_column",
    "STG1_Type": "stg1_type", "STG1_Length": "stg1_length", "STG1_Precision": "stg1_precision",
    "SRC_Source_Table": "src_source_table", "SRC_Source_Column": "src_source_column",
    "SRC_to_STG1_Transformation": "src_to_stg1_transform",
    "STG1_to_STG2_Transformation": "stg1_to_stg2_transform",
    "Lineage_Status": "lineage_status", "Lineage_Status_Detail": "lineage_status_detail",
}


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _headers(ws):
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [(_s(c) or "") for c in row]
    return []


def _parse_ud(blob):
    """Return dict of UD_n -> value, or {} if not parseable."""
    if not blob:
        return {}
    try:
        d = json.loads(blob)
        return {k: (str(v) if v is not None else "") for k, v in d.items()} if isinstance(d, dict) else {}
    except Exception:
        return {}


class LegacyLineageConnector:
    name = "legacy_lineage"

    def __init__(self, xlsx_path, lineage_sheet=None, proof_sheet=None):
        self.xlsx_path = xlsx_path
        self.lineage_sheet = lineage_sheet
        self.proof_sheet = proof_sheet

    @classmethod
    def from_env(cls):
        return cls(
            os.environ.get("CP_LEGACY_LINEAGE_XLSX", "sample-artifacts/LEGACY-LINEAGE/legacy_lineage.xlsx"),
            os.environ.get("CP_LEGACY_LINEAGE_SHEET"),
            os.environ.get("CP_LEGACY_PROOF_SHEET"),
        )

    def parse(self):
        if not os.path.exists(self.xlsx_path):
            log.warning("legacy lineage workbook not found: %s (skipping)", self.xlsx_path)
            return {"lineage": [], "proof": []}
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        lineage = self._parse_lineage(wb)
        proof = self._parse_proof(wb)
        log.info("legacy_lineage: %d lineage rows, %d proof rows", len(lineage), len(proof))
        return {"lineage": lineage, "proof": proof}

    # ---- lineage sheet ----
    def _parse_lineage(self, wb):
        name = self.lineage_sheet
        if not name:
            for s in wb.sheetnames:
                low = s.lower().replace(" ", "")
                if "lineage" in low or "datalineage" in low or "schema" in low:
                    name = s
                    break
        if not name:
            name = wb.sheetnames[0]
        ws = wb[name]
        hdr = _headers(ws)
        idx = {h: i for i, h in enumerate(hdr)}
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(col):
                i = idx.get(col)
                return _s(row[i]) if i is not None and i < len(row) else None
            dwh_t, dwh_c = g("DWH_Target_Table"), g("DWH_Target_Column")
            if not dwh_t or not dwh_c:
                continue
            rec = {v: g(k) for k, v in LINEAGE_MAP.items()}
            rec["lineage_id"] = f"{dwh_t}:{dwh_c}"
            rec["is_ud"] = "N"
            rec["ud_key"] = None
            out.append(rec)
        return out

    # ---- proof sheet (also drives UD explosion, since values live here) ----
    def _parse_proof(self, wb):
        # the proof/variance sheet is named "e2e" in the BBH workbook; fall back
        # to other common names. Must NOT match the lineage sheet.
        name = self.proof_sheet
        if not name:
            for s in wb.sheetnames:
                low = s.lower().replace(" ", "")
                if low in ("e2e", "proof", "variance") or "e2e" in low or "proof" in low:
                    name = s
                    break
        if not name or name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []
        # layout: row for "Table"/table name, a "Headers" row listing fields,
        # then one row per stage (DWH/STG2/STG1/SRC) with aligned values.
        table_name = None
        header_fields = []
        stage_rows = []
        for r in rows:
            cells = [_s(c) for c in r]
            if not any(cells):
                continue
            tag = (cells[0] or "").upper()
            if tag == "TABLE":
                table_name = next((c for c in cells[1:] if c), None)
            elif tag == "HEADERS":
                header_fields = [c for c in cells[1:] if c]
            elif tag in ("DWH", "STG2", "STG1", "SRC"):
                stage_rows.append((tag, cells[1:]))
        if not header_fields:
            return []
        out = []
        for stage, vals in stage_rows:
            for i, field in enumerate(header_fields):
                val = vals[i] if i < len(vals) else None
                pid = f"{table_name}:{field}:{stage}"
                out.append({"proof_id": pid, "proof_table": table_name, "field_name": field,
                            "stage": stage, "field_value": val, "is_ud": "N", "ud_key": None})
                # explode UD CLOB values into per-key proof rows (keep original too)
                if field and field.upper() == UD_CLOB_COL:
                    for k, v in _parse_ud(val).items():
                        out.append({"proof_id": f"{table_name}:{field}:{stage}:{k}",
                                    "proof_table": table_name, "field_name": k, "stage": stage,
                                    "field_value": v, "is_ud": "Y", "ud_key": k})
        return out

    @staticmethod
    def _guess(wb, needles):
        for s in wb.sheetnames:
            low = s.lower().replace(" ", "")
            if any(n in low for n in needles):
                return s
        return None

    def load(self, loader, bundle):
        for r in bundle.get("lineage", []):
            loader._merge("legacy_lineage", ("lineage_id",), r)
        for p in bundle.get("proof", []):
            loader._merge("legacy_proof", ("proof_id",), p)
        # build UD lineage rows from proof UD keys that lack a lineage entry,
        # so exploded UD fields still appear in the backward trace.
        seen = {r["lineage_id"] for r in bundle.get("lineage", [])}
        for p in bundle.get("proof", []):
            if p.get("is_ud") == "Y":
                lid = f"{p['proof_table']}:{p['field_name']}"
                if lid not in seen:
                    seen.add(lid)
                    loader._merge("legacy_lineage", ("lineage_id",), {
                        "lineage_id": lid, "dwh_target_table": p["proof_table"],
                        "dwh_target_column": p["field_name"], "is_ud": "Y", "ud_key": p["ud_key"],
                        "lineage_status": "UD Attribute",
                        "lineage_status_detail": "Exploded from USER_DEFINED_ATTRIBUTE_CLOB"})
